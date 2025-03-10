from datetime import date
from functools import cached_property
from flask import current_app
from lbrc_flask.database import db
from lbrc_flask.security import AuditMixin
from lbrc_flask.model import CommonMixin
from lbrc_flask.validators import is_integer, parse_date_or_none
from sqlalchemy.orm import Mapped, mapped_column, relationship
from sqlalchemy import ForeignKey, String, Text
from werkzeug.utils import secure_filename
from itertools import takewhile
from openpyxl import load_workbook

class Upload(AuditMixin, CommonMixin, db.Model):
    COLUMNS = {
        'key': dict(type='int', allow_null=True),
        'freezer': dict(type='int'),
        'drawer': dict(type='int'),
        'box_number': dict(type='int'),
        'position': dict(type='str', max_length=20),
        'bacterial species': dict(type='str', max_length=100),
        'strain': dict(type='str', max_length=100),
        'media': dict(type='str', max_length=100),
        'plasmid name': dict(type='str', max_length=100),
        'resistance marker': dict(type='str', max_length=100),
        'phage id': dict(type='str', max_length=100),
        'host species': dict(type='str', max_length=100),
        'description': dict(type='str'),
        'project': dict(type='str', max_length=100),
        'date': dict(type='date'),
        'storage method': dict(type='str', max_length=100),
        'name': dict(type='str'),
        'notes': dict(type='str'),
    }

    STATUS__AWAITING_PROCESSING = 'Awaiting Processing'
    STATUS__PROCESSED = 'Processed'
    STATUS__ERROR = 'Error'

    STATUS_NAMES = [
        STATUS__AWAITING_PROCESSING,
        STATUS__PROCESSED,
        STATUS__ERROR,
    ]

    COLUMN_NAMES = set(COLUMNS.keys())

    id: Mapped[int] = mapped_column(primary_key=True)
    filename: Mapped[str] = mapped_column(String(500))
    status: Mapped[str] = mapped_column(String(50), default='')
    errors: Mapped[str] = mapped_column(Text, default='')

    @property
    def local_filepath(self):
        return current_app.config["FILE_UPLOAD_DIRECTORY"] / secure_filename(f"{self.id}_{self.filename}")

    def worksheet(self):
        wb = load_workbook(filename=self.local_filepath, read_only=True)
        return wb.active

    @cached_property
    def column_names(self):
        rows = self.worksheet().iter_rows(min_row=1, max_row=1)
        first_row = next(rows)

        result = [c.value.lower() for c in takewhile(lambda x: x.value, first_row)]

        return result

    def iter_rows(self):
        for r in self.worksheet().iter_rows(values_only=True):
            values = dict(zip(self.column_names, r))
            yield values

    def iter_data(self):
        for r in self.iter_rows:
            if self._is_data_row(r):
                yield r

    def _is_data_row(self, row):
        return 'key' in row.keys() and (row.get('key', None) is None or is_integer(row['key']))

    def validate(self):
        errors = []

        errors.extend(self._column_validation_errors())
        errors.extend(self._data_validation_errors())

        if errors:
            self.errors = "\n".join(errors)
            self.status = Upload.STATUS__ERROR
    
    def _column_validation_errors(self):
        missing_columns = Upload.COLUMN_NAMES - set(self.column_names)
        return map(lambda x: f"Missing column '{x}'", missing_columns)

    def _data_validation_errors(self):
        errors = []

        for i, row in enumerate(self.iter_rows):
            if self._is_ambigous_row(row):
                errors.append(f"Row {i}: contains columns for both bacteria and phages")
            elif self._neither_phage_nor_bacterium(row):
                errors.append(f"Row {i}: does not contain enough information")
            elif errors := self._bacterium_errors(row):
                for e in errors:
                    errors.append(f"Row {i}: {e}")
            elif errors := self._phage_errors(row):
                for e in errors:
                    errors.append(f"Row {i}: {e}")

        return errors

    def _is_invalid_string(self, value, column, col_def):
        max_length = col_def.get('max_length', None)
        
        if not max_length:
            return
        
        if value is None or len(value) > max_length:
            return f"Text too long in column '{column}'"

    def _is_invalid_interger(self, value, column, col_def):
        if value is None or not is_integer(value):
            return f"Invalid value in column '{column}'"

    def _is_invalid_date(self, value, column, col_def):
        if parse_date_or_none(value) is None:
            return f"Invalid value in column '{column}'"

    @property
    def is_error(self):
        return self.status == Upload.STATUS__ERROR


class Lookup(AuditMixin, CommonMixin):
    id: Mapped[int] = mapped_column(primary_key=True)
    name: Mapped[str] = mapped_column(String(100), index=True, unique=True)

    def __str__(self):
        return self.name

class BacterialSpecies(Lookup, db.Model):
    pass

class Strain(Lookup, db.Model):
    pass

class Medium(Lookup, db.Model):
    pass

class Plasmid(Lookup, db.Model):
    pass

class ResistanceMarker(Lookup, db.Model):
    pass

class PhageIdentifier(Lookup, db.Model):
    pass

class Project(Lookup, db.Model):
    pass

class StorageMethod(Lookup, db.Model):
    pass

class StaffMember(Lookup, db.Model):
    pass

class Specimen(AuditMixin, CommonMixin, db.Model):
    id: Mapped[int] = mapped_column(primary_key=True)
    type: Mapped[str] = mapped_column(String(20), index=True)

    __mapper_args__ = {
        "polymorphic_on": type,
    }

    freezer: Mapped[int] = mapped_column(index=True)
    draw: Mapped[int] = mapped_column(index=True)
    position: Mapped[str] = mapped_column(String(20), index=True)
    name: Mapped[str] = mapped_column(Text, index=True)
    description: Mapped[str] = mapped_column(Text, index=True)
    notes: Mapped[str] = mapped_column(Text, index=True)
    sample_date: Mapped[date] = mapped_column(index=True)

    project_id: Mapped[int] = mapped_column(ForeignKey(Project.id), index=True, nullable=True)
    project: Mapped[Project] = relationship(foreign_keys=[project_id])
    storage_method_id: Mapped[int] = mapped_column(ForeignKey(StorageMethod.id), index=True, nullable=True)
    storage_method: Mapped[StorageMethod] = relationship(foreign_keys=[storage_method_id])
    staff_member_id: Mapped[int] = mapped_column(ForeignKey(StaffMember.id), index=True, nullable=True)
    staff_member: Mapped[StaffMember] = relationship(foreign_keys=[staff_member_id])

    @property
    def is_bacterium(self):
        return False

    @property
    def is_phage(self):
        return False

    def data(self):
        return dict(
            key=self.id,
            freezer=self.freezer,
            draw=self.draw,
            position=self.position,
            name=self.name,
            description=self.description,
            notes=self.notes,
            sample_date=self.sample_date,
            project=self.project.name,
            storage_method=self.storage_method.name,
            staff_member=self.staff_member.name,
        )

class Bacterium(Specimen):
    __mapper_args__ = {
        "polymorphic_identity": 'Bacterium',
    }

    species_id: Mapped[int] = mapped_column(ForeignKey(BacterialSpecies.id), index=True, nullable=True)
    species: Mapped[BacterialSpecies] = relationship(foreign_keys=[species_id])
    strain_id: Mapped[int] = mapped_column(ForeignKey(Strain.id), index=True, nullable=True)
    strain: Mapped[Strain] = relationship(foreign_keys=[strain_id])
    medium_id: Mapped[int] = mapped_column(ForeignKey(Medium.id), index=True, nullable=True)
    medium: Mapped[Medium] = relationship(foreign_keys=[medium_id])
    plasmid_id: Mapped[int] = mapped_column(ForeignKey(Plasmid.id), index=True, nullable=True)
    plasmid: Mapped[Plasmid] = relationship(foreign_keys=[plasmid_id])
    resistance_marker_id: Mapped[int] = mapped_column(ForeignKey(ResistanceMarker.id), index=True, nullable=True)
    resistance_marker: Mapped[ResistanceMarker] = relationship(foreign_keys=[resistance_marker_id])

    @property
    def is_bacterium(self):
        return True
    
    def data(self):
        result = super().data()

        result['species'] = self.species.name
        result['strain'] = self.strain.name
        result['medium'] = self.medium.name
        result['plasmid'] = self.plasmid.name
        result['resistance_marker'] = self.resistance_marker.name

        return result


class Phage(Specimen):
    __mapper_args__ = {
        "polymorphic_identity": 'Phage',
    }

    phage_identifier_id: Mapped[int] = mapped_column(ForeignKey(PhageIdentifier.id), index=True, nullable=True)
    phage_identifier: Mapped[PhageIdentifier] = relationship(foreign_keys=[phage_identifier_id])
    host_id: Mapped[int] = mapped_column(ForeignKey(BacterialSpecies.id), index=True, nullable=True)
    host: Mapped[BacterialSpecies] = relationship(foreign_keys=[host_id])

    @property
    def is_phage(self):
        return True

    def data(self):
        result = super().data()

        result['phage_identifier'] = self.phage_identifier.name
        result['host'] = self.host.name

        return result
