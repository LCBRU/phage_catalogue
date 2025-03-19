from copy import deepcopy
from dataclasses import dataclass
from functools import cached_property
from pathlib import Path
from typing import Optional
from flask import current_app
from lbrc_flask.database import db
from lbrc_flask.security import AuditMixin
from lbrc_flask.model import CommonMixin
from lbrc_flask.validators import is_integer, parse_date_or_none
from sqlalchemy.orm import Mapped, mapped_column
from sqlalchemy import String, Text
from werkzeug.utils import secure_filename
from itertools import compress, islice, takewhile
from openpyxl import load_workbook

from phage_catalogue.model.specimens import Bacterium


class Upload(AuditMixin, CommonMixin, db.Model):
    STATUS__AWAITING_PROCESSING = 'Awaiting Processing'
    STATUS__PROCESSED = 'Processed'
    STATUS__ERROR = 'Error'

    STATUS_NAMES = [
        STATUS__AWAITING_PROCESSING,
        STATUS__PROCESSED,
        STATUS__ERROR,
    ]

    id: Mapped[int] = mapped_column(primary_key=True)
    filename: Mapped[str] = mapped_column(String(500))
    status: Mapped[str] = mapped_column(String(50), default='')
    errors: Mapped[str] = mapped_column(Text, default='')

    @property
    def local_filepath(self):
        return current_app.config["FILE_UPLOAD_DIRECTORY"] / secure_filename(f"{self.id}_{self.filename}")

    def validate(self):
        spreadsheet = Spreadsheet(self.local_filepath)
        upload_column_definition = UploadColumnDefinition()

        errors = upload_column_definition.validation_errors(spreadsheet)

        if errors:
            self.errors = "\n".join(errors)
            self.status = Upload.STATUS__ERROR

    @property
    def is_error(self):
        return self.status == Upload.STATUS__ERROR

    def bacteria_data(self):
        spreadsheet = BacteriaFullSpreadsheetDefinition(Spreadsheet(self.local_filepath))

        return spreadsheet.translated_data()

    def phages_data(self):
        spreadsheet = PhageFullSpreadsheetDefinition(Spreadsheet(self.local_filepath))

        return spreadsheet.translated_data()


class ColumnsDefinition():
    @property
    def column_definition(self):
        return []
    
    @property
    def column_names(self):
        return [d.name for d in self.column_definition]

    def definition_for_column_name(self, name):
        for d in self.column_definition:
            if d.name == name:
                return d


@dataclass
class ColumnDefinition:
    COLUMN_TYPE_STRING = 'str'
    COLUMN_TYPE_INTEGER = 'int'
    COLUMN_TYPE_DATE = 'date'

    name: str
    type: str
    allow_null: bool = False
    max_length: Optional[int] = None
    translated_name: Optional[str] = None

    def value(self, row: dict):
        return row.get(self.name, None)

    def stringed_value(self, row: dict):
        return str(self.value(row) or '').strip()

    def has_value(self, row: dict):
        return len(self.stringed_value(row)) > 0

    def get_translated_name(self):
        return self.translated_name or self.name


class UploadColumnDefinition(ColumnsDefinition):
    @property
    def column_definition(self):
        result = deepcopy(SpecimenColumnDefinition().column_definition)
        result.extend(BacteriumOnlyColumnDefinition().column_definition)
        result.extend(PhageOnlyColumnDefinition().column_definition)

        return result

    def validation_errors(self, spreadsheet):
        errors = set()

        upload_spreadsheet = SpreadsheetDefinition(spreadsheet, self)
        bactetria_spreadsheet = BacteriaFullSpreadsheetDefinition(spreadsheet)
        phage_spreadsheet = PhageFullSpreadsheetDefinition(spreadsheet)

        errors = errors.union(upload_spreadsheet.column_validation_errors())
        errors = errors.union(self._both_phage_and_bacterium_errors(spreadsheet))
        errors = errors.union(self._not_enough_columns_errors(spreadsheet))
        errors = errors.union(bactetria_spreadsheet.data_validation_errors())
        errors = errors.union(phage_spreadsheet.data_validation_errors())

        return errors

    def _not_enough_columns_errors(self, spreadsheet):
        result = []

        bactetria_full_spreadsheet = BacteriaFullSpreadsheetDefinition(spreadsheet)
        phage_full_spreadsheet = PhageFullSpreadsheetDefinition(spreadsheet)

        for i, (baceterium, phage) in enumerate(zip(bactetria_full_spreadsheet.rows_with_all_fields, phage_full_spreadsheet.rows_with_all_fields), 1):
            if not(phage or baceterium):
                result.append(f"Row {i}: does not contain enough information")

        return result

    def _both_phage_and_bacterium_errors(self, spreadsheet):
        result = []

        bactetria_only_spreadsheet = BacteriaOnlySpreadsheetDefinition(spreadsheet)
        phage_only_spreadsheet = PhageOnlySpreadsheetDefinition(spreadsheet)

        for i, (bacterium, phage) in enumerate(zip(bactetria_only_spreadsheet.rows_with_any_fields, phage_only_spreadsheet.rows_with_any_fields), 1):
            if bacterium and phage:
                result.append(f"Row {i}: contains columns for both bacteria and phages")
        
        return result
    

class SpecimenColumnDefinition(ColumnsDefinition):
    @property
    def column_definition(self):
        return [
            ColumnDefinition(
                name='key',
                type=ColumnDefinition.COLUMN_TYPE_INTEGER,
                allow_null=True,
            ),
            ColumnDefinition(
                name='freezer',
                type=ColumnDefinition.COLUMN_TYPE_INTEGER,
            ),
            ColumnDefinition(
                name='drawer',
                type=ColumnDefinition.COLUMN_TYPE_INTEGER,
            ),
            ColumnDefinition(
                name='box_number',
                type=ColumnDefinition.COLUMN_TYPE_STRING,
                max_length=100,
            ),
            ColumnDefinition(
                name='position',
                type=ColumnDefinition.COLUMN_TYPE_STRING,
                max_length=20,
            ),
            ColumnDefinition(
                name='description',
                type=ColumnDefinition.COLUMN_TYPE_STRING,
            ),
            ColumnDefinition(
                name='project',
                type=ColumnDefinition.COLUMN_TYPE_STRING,
                max_length=100,
            ),
            ColumnDefinition(
                name='date',
                type=ColumnDefinition.COLUMN_TYPE_DATE,
                translated_name='sample_date',
            ),
            ColumnDefinition(
                name='storage method',
                type=ColumnDefinition.COLUMN_TYPE_STRING,
                max_length=100,
                translated_name='storage_method',
            ),
            ColumnDefinition(
                name='name',
                type=ColumnDefinition.COLUMN_TYPE_STRING,
            ),
            ColumnDefinition(
                name='staff member',
                type=ColumnDefinition.COLUMN_TYPE_STRING,
                translated_name='staff_member',
            ),
            ColumnDefinition(
                name='notes',
                type=ColumnDefinition.COLUMN_TYPE_STRING,
            ),
        ]


class BacteriumOnlyColumnDefinition(ColumnsDefinition):
    @property
    def column_definition(self):
        return [
            ColumnDefinition(
                name='bacterial species',
                type=ColumnDefinition.COLUMN_TYPE_STRING,
                max_length=100,
                translated_name='species',
            ),
            ColumnDefinition(
                name='strain',
                type=ColumnDefinition.COLUMN_TYPE_STRING,
                max_length=100,
            ),
            ColumnDefinition(
                name='media',
                type=ColumnDefinition.COLUMN_TYPE_STRING,
                max_length=100,
                translated_name='medium',
            ),
            ColumnDefinition(
                name='plasmid name',
                type=ColumnDefinition.COLUMN_TYPE_STRING,
                max_length=100,
                translated_name='plasmid',
            ),
            ColumnDefinition(
                name='resistance marker',
                type=ColumnDefinition.COLUMN_TYPE_STRING,
                max_length=100,
                translated_name='resistance_marker',
            ),
        ]


class BacteriumFullColumnDefinition(ColumnsDefinition):
    @property
    def column_definition(self):
        result = deepcopy(SpecimenColumnDefinition().column_definition)
        result.extend(BacteriumOnlyColumnDefinition().column_definition)

        return result

class PhageOnlyColumnDefinition(ColumnsDefinition):
    @property
    def column_definition(self):
        return [
            ColumnDefinition(
                name='phage id',
                type=ColumnDefinition.COLUMN_TYPE_STRING,
                max_length=100,
                translated_name='phage_identifier',
            ),
            ColumnDefinition(
                name='host species',
                type=ColumnDefinition.COLUMN_TYPE_STRING,
                max_length=100,
                translated_name='host',
            ),
        ]


class PhageFullColumnDefinition(ColumnsDefinition):
    @property
    def column_definition(self):
        result = deepcopy(SpecimenColumnDefinition().column_definition)
        result.extend(PhageOnlyColumnDefinition().column_definition)

        return result


class Spreadsheet():
    def __init__(self, filepath: Path, header_rows: int = 1):
        self.filepath: Path = filepath
        self.header_rows: int = header_rows

    def worksheet(self):
        wb = load_workbook(filename=self.filepath, read_only=True)
        return wb.active

    @cached_property
    def column_names(self):
        rows = self.worksheet().values
        first_row = next(rows)

        result = [c.lower() for c in takewhile(lambda x: x, first_row)]

        return result

    def iter_rows(self):
        for r in self.worksheet().values:
            yield dict(zip(self.column_names, r))

    def iter_data(self):
        for r in islice(self.iter_rows(), self.header_rows, None):
            yield r


class SpreadsheetDefinition:
    def __init__(self, spreadsheet: Spreadsheet, column_definition: ColumnsDefinition):
        self.spreadsheet: Spreadsheet = spreadsheet
        self.column_definition: ColumnsDefinition = column_definition
    
    def validation_errors(self):
        errors = []

        errors.extend(self.column_validation_errors())
        errors.extend(self.data_validation_errors())

        return errors

    def column_validation_errors(self):
        missing_columns = set(self.column_definition.column_names) - set(self.spreadsheet.column_names)
        return map(lambda x: f"Missing column '{x}'", missing_columns)

    def iter_filtered_data(self):
        return compress(self.spreadsheet.iter_data(), self.rows_with_all_fields)
    
    def translated_data(self):
        for row in self.iter_filtered_data():
            result = {}

            for cd in self.column_definition.column_definition:
                result[cd.get_translated_name()] = cd.value(row)
            
            yield result

    def data_validation_errors(self):
        result = []

        for i, row in enumerate(self.iter_filtered_data(), 1):
            row_errors = self._field_errors_for_def(row)
            result.extend(map(lambda e: f"Row {i}: {e}", row_errors))

        return result
    
    @cached_property
    def rows_with_all_fields(self):
        result = []

        for row in self.spreadsheet.iter_data():
            result.append(all([d.has_value(row) or d.allow_null for d in self.column_definition.column_definition]))

        return result

    @cached_property
    def rows_with_any_fields(self):
        result = []

        for row in self.spreadsheet.iter_data():
            result.append(any([d.has_value(row) for d in self.column_definition.column_definition]))

        return result

    def _field_errors_for_def(self, row: dict):
        result = []
        for col_def in self.column_definition.column_definition:
            if col_def.name in row:
                result.extend(self._field_errors(row, col_def))
        
        return result
    
    def _field_errors(self, row, col_def):
        result = []

        value = row[col_def.name]

        if not col_def.allow_null:
            is_null = value is None or str(value).strip() == ''
            if is_null:
                result.append("Data is mising")

        match col_def.type:
            case ColumnDefinition.COLUMN_TYPE_STRING:
                result.extend(self._is_invalid_string(value, col_def))
            case ColumnDefinition.COLUMN_TYPE_INTEGER:
                result.extend(self._is_invalid_interger(value, col_def))
            case ColumnDefinition.COLUMN_TYPE_DATE:
                result.extend(self._is_invalid_date(value, col_def))
        
        return map(lambda e: f"{col_def.name}: {e}", result)

    def _is_invalid_string(self, value, col_def):
        if value is None:
            return []

        if max_length := col_def.max_length:
            if len(value) > max_length:
               return [f"Text is longer than {max_length} characters"]

        return []

    def _is_invalid_interger(self, value, col_def):
        if value is None:
            return []

        if not is_integer(value):
            return ["Invalid value"]
        
        return []

    def _is_invalid_date(self, value, col_def):
        if value is None:
            return []

        if parse_date_or_none(value) is None:
            return ["Invalid value"]
        
        return []


class BacteriaFullSpreadsheetDefinition(SpreadsheetDefinition):
    def __init__(self, spreadsheet):
        super().__init__(spreadsheet, BacteriumFullColumnDefinition())
    
class BacteriaOnlySpreadsheetDefinition(SpreadsheetDefinition):
    def __init__(self, spreadsheet):
        super().__init__(spreadsheet, BacteriumOnlyColumnDefinition())


class PhageFullSpreadsheetDefinition(SpreadsheetDefinition):
    def __init__(self, spreadsheet):
        super().__init__(spreadsheet, PhageFullColumnDefinition())


class PhageOnlySpreadsheetDefinition(SpreadsheetDefinition):
    def __init__(self, spreadsheet):
        super().__init__(spreadsheet, PhageOnlyColumnDefinition())

